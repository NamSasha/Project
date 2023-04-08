import os 
import openpyxl
from openpyxl import load_workbook
from tkinter import filedialog, Tk
import time
from tqdm import tqdm
import zipfile
from wand.image import Image
import shutil



def process(docx, img_dir=None):
    text = u''

    # unzip the docx in memory
    zipf = zipfile.ZipFile(docx)
    filelist = zipf.namelist()

    if img_dir is not None:
        # extract images
        for fname in filelist:
            _, extension = os.path.splitext(fname)
            if extension in [".jpg", ".jpeg", ".png", ".bmp",".emf"]:
                dst_fname = os.path.join(img_dir, os.path.basename(fname))
                with open(dst_fname, "wb") as dst_f:
                    dst_f.write(zipf.read(fname))
                    
    zipf.close()
    return text.strip()

#headers
print("Hi, Sasha here. Greetinngss.......\n\n")
for i in tqdm(range(0,10),desc= 'Đợi 1 chút :3'):
    time.sleep(0.1)
Tk().withdraw()


#load folder contains word
foldername = filedialog.askdirectory(title="Choose Force log folder")  
start_time = time.time()
files = os.listdir(foldername) 

#make new result path
resultpath = os.path.join(foldername, 'result')    
os.makedirs(resultpath) 

#convert image from docx to emf to png format
for file in files:
    file_name = file.replace('.docx','')
    path = foldername + "/" + file
    process(path,resultpath)
    raw = Image(filename = resultpath + '/' + 'image1.emf')
    raw_convert = raw.convert('png')
    raw_convert.save(filename = resultpath + '/' + file_name + '.png')
    

#delete emf file
os.remove(resultpath +'/' + 'image1.emf')
img_names = os.listdir(resultpath)

#load excel
excel_path = filedialog.askopenfilename(title = "Select excel file",filetypes = (("Xlsm Files","*.xlsm"),("All","*.*")))
wb = load_workbook(excel_path,keep_vba=True)
ws = wb['LCB']


img_names = os.listdir(resultpath)
#add_image
for image in img_names:
    img = image.replace('.png','')
    image_path = resultpath +'/'+ image
    mycell_img = ws.cell(row = int(img)+32,column = 13).coordinate
    imgs = openpyxl.drawing.image.Image(image_path)
    imgs.width = 175
    imgs.height = 105
    ws.add_image(imgs,mycell_img) 
wb.save(excel_path)

shutil.rmtree(resultpath)