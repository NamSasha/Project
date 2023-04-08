import glob
import os 
import openpyxl
from openpyxl import load_workbook
from tkinter import END, filedialog
from win32com.client import Dispatch
import time
from tqdm import tqdm


#headers
print("Hi, Sasha here. Greetinngss.......\n\n")
for i in tqdm(range(0,10),desc= 'Đợi 1 chút :3'):
    time.sleep(0.1)

#directory
print('Choose Image Directory')
img_path = filedialog.askdirectory(title='Select Image Directory')
print('Choose Excel_path')
excel_path = filedialog.askopenfilename(title = "Select excel file",filetypes = (("Xlsm Files","*.xlsm"),("All","*.*")))

#get data
for x in glob.glob(img_path):
    img_names = os.listdir(x)
    
#img_names,barcode,DP
barcode = [barcode.split('_')[0] for barcode in img_names]
DP = [dp.split('_')[1].split('.')[0] for dp in img_names]
img_names =  [img_path+'/'+ img for img in img_names]
#modifiy excel
wb = load_workbook(excel_path,keep_vba=True)
ws = wb['DP']

#delete old value
for row in ws['B33:F204']:
    for cell in row:
        cell.value = None
while ws._images:
    del ws._images[0]

#insert new value
for index,rows in zip(range(1,len(barcode)+1),range(33,len(barcode)+34)):
    mycell_barcode =  ws.cell(row = rows,column = 5)
    mycell_DP = ws.cell(row = rows,column = 6)
    mycell_index = ws.cell(row = rows,column = 2)
    mycell_barcode.value = barcode[rows-33]
    mycell_DP.value = DP[rows-33]
    mycell_index.value = index

#insert image
for image,rows in zip(img_names,range(33,len(barcode)+34)):
    mycell_img =  ws.cell(row = rows,column = 7).coordinate
    img = openpyxl.drawing.image.Image(image)
    img.width = 180
    img.height = 110
    ws.add_image(img,mycell_img)

#save        
wb.save(excel_path)
